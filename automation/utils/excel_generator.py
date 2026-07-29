import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import EXCEL_REPORTS_DIR
from automation.utils.logger import get_logger

logger = get_logger("ExcelReportGenerator")

def style_header_cell(cell, text, fill_color="1F4E79", font_color="FFFFFF"):
    cell.value = text
    cell.font = Font(name="Calibri", size=11, bold=True, color=font_color)
    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def apply_borders(ws):
    thin = Side(border_style="thin", color="D9D9D9")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

def auto_fit_columns(ws):
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def generate_excel_reports(results, summary_metrics):
    """
    Generates:
    1. Automation_Test_Report.xlsx (6 Sheets)
    2. Failed_Test_Cases.xlsx
    3. Passed_Test_Cases.xlsx
    4. Summary_Report.xlsx
    """
    logger.info("Generating Excel reports...")
    
    # ----------------------------------------------------
    # 1. Automation_Test_Report.xlsx (6 Sheets)
    # ----------------------------------------------------
    wb_main = openpyxl.Workbook()
    wb_main.remove(wb_main.active) # Remove default sheet
    
    # Sheet 1: Executed Test Cases
    ws_all = wb_main.create_sheet(title="Executed Test Cases")
    headers_all = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority", "Failure Reason"]
    for col_idx, h in enumerate(headers_all, 1):
        style_header_cell(ws_all.cell(row=1, column=col_idx), h)
        
    row_idx = 2
    for r in results:
        ws_all.cell(row=row_idx, column=1, value=r["id"])
        ws_all.cell(row=row_idx, column=2, value=r["module"])
        ws_all.cell(row=row_idx, column=3, value=r["name"])
        
        status_cell = ws_all.cell(row=row_idx, column=4, value=r["status"])
        if r["status"] == "PASS":
            status_cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
            status_cell.font = Font(color="006100", bold=True)
        elif r["status"] == "FAIL":
            status_cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")
            status_cell.font = Font(color="9C0006", bold=True)
        else:
            status_cell.fill = PatternFill(start_color="FFEB9C", fill_type="solid")
            status_cell.font = Font(color="9C6500", bold=True)
            
        ws_all.cell(row=row_idx, column=5, value=r.get("execution_time", 0.05))
        ws_all.cell(row=row_idx, column=6, value=r.get("priority", "P2"))
        ws_all.cell(row=row_idx, column=7, value=r.get("failure_reason", ""))
        row_idx += 1
    apply_borders(ws_all)
    auto_fit_columns(ws_all)

    # Sheet 2: Passed Tests
    ws_pass = wb_main.create_sheet(title="Passed Tests")
    for col_idx, h in enumerate(headers_all[:6], 1):
        style_header_cell(ws_pass.cell(row=1, column=col_idx), h, fill_color="27AE60")
    p_row = 2
    for r in results:
        if r["status"] == "PASS":
            ws_pass.cell(row=p_row, column=1, value=r["id"])
            ws_pass.cell(row=p_row, column=2, value=r["module"])
            ws_pass.cell(row=p_row, column=3, value=r["name"])
            ws_pass.cell(row=p_row, column=4, value="PASS")
            ws_pass.cell(row=p_row, column=5, value=r.get("execution_time", 0.05))
            ws_pass.cell(row=p_row, column=6, value=r.get("priority", "P2"))
            p_row += 1
    apply_borders(ws_pass)
    auto_fit_columns(ws_pass)

    # Sheet 3: Failed Tests
    ws_fail = wb_main.create_sheet(title="Failed Tests")
    for col_idx, h in enumerate(headers_all, 1):
        style_header_cell(ws_fail.cell(row=1, column=col_idx), h, fill_color="C0392B")
    f_row = 2
    for r in results:
        if r["status"] == "FAIL":
            ws_fail.cell(row=f_row, column=1, value=r["id"])
            ws_fail.cell(row=f_row, column=2, value=r["module"])
            ws_fail.cell(row=f_row, column=3, value=r["name"])
            ws_fail.cell(row=f_row, column=4, value="FAIL")
            ws_fail.cell(row=f_row, column=5, value=r.get("execution_time", 0.05))
            ws_fail.cell(row=f_row, column=6, value=r.get("priority", "P2"))
            ws_fail.cell(row=f_row, column=7, value=r.get("failure_reason", ""))
            f_row += 1
    apply_borders(ws_fail)
    auto_fit_columns(ws_fail)

    # Sheet 4: Skipped Tests
    ws_skip = wb_main.create_sheet(title="Skipped Tests")
    for col_idx, h in enumerate(headers_all[:6], 1):
        style_header_cell(ws_skip.cell(row=1, column=col_idx), h, fill_color="D35400")
    s_row = 2
    for r in results:
        if r["status"] == "SKIP":
            ws_skip.cell(row=s_row, column=1, value=r["id"])
            ws_skip.cell(row=s_row, column=2, value=r["module"])
            ws_skip.cell(row=s_row, column=3, value=r["name"])
            ws_skip.cell(row=s_row, column=4, value="SKIP")
            ws_skip.cell(row=s_row, column=5, value=r.get("execution_time", 0.0))
            ws_skip.cell(row=s_row, column=6, value=r.get("priority", "P2"))
            s_row += 1
    apply_borders(ws_skip)
    auto_fit_columns(ws_skip)

    # Sheet 5: Execution Metrics
    ws_metrics = wb_main.create_sheet(title="Execution Metrics")
    metrics_headers = ["Metric Name", "Metric Value"]
    for col_idx, h in enumerate(metrics_headers, 1):
        style_header_cell(ws_metrics.cell(row=1, column=col_idx), h, fill_color="2980B9")
    m_data = [
        ("Total Test Cases", summary_metrics["total"]),
        ("Passed Tests", summary_metrics["passed"]),
        ("Failed Tests", summary_metrics["failed"]),
        ("Skipped Tests", summary_metrics["skipped"]),
        ("Pass Rate (%)", f"{summary_metrics['pass_rate']:.2f}%"),
        ("Execution Duration (s)", f"{summary_metrics['duration']:.2f}s"),
        ("Environment / Base URL", summary_metrics.get("base_url", "GitHub Pages"))
    ]
    for idx, (k, v) in enumerate(m_data, 2):
        ws_metrics.cell(row=idx, column=1, value=k)
        ws_metrics.cell(row=idx, column=2, value=v)
    apply_borders(ws_metrics)
    auto_fit_columns(ws_metrics)

    # Sheet 6: Defect Summary
    ws_defects = wb_main.create_sheet(title="Defect Summary")
    defect_headers = ["Defect ID", "Module", "Test ID", "Failure Summary", "Severity", "Status"]
    for col_idx, h in enumerate(defect_headers, 1):
        style_header_cell(ws_defects.cell(row=1, column=col_idx), h, fill_color="8E44AD")
    d_row = 2
    for r in results:
        if r["status"] == "FAIL":
            ws_defects.cell(row=d_row, column=1, value=f"DEF-{r['id']}")
            ws_defects.cell(row=d_row, column=2, value=r["module"])
            ws_defects.cell(row=d_row, column=3, value=r["id"])
            ws_defects.cell(row=d_row, column=4, value=r.get("failure_reason", "Assertion Failed"))
            ws_defects.cell(row=d_row, column=5, value="High" if r.get("priority") == "P1" else "Medium")
            ws_defects.cell(row=d_row, column=6, value="OPEN")
            d_row += 1
    apply_borders(ws_defects)
    auto_fit_columns(ws_defects)

    main_excel_path = EXCEL_REPORTS_DIR / "Automation_Test_Report.xlsx"
    wb_main.save(main_excel_path)
    logger.info(f"Saved {main_excel_path}")

    # ----------------------------------------------------
    # 2. Failed_Test_Cases.xlsx
    # ----------------------------------------------------
    wb_failed = openpyxl.Workbook()
    ws_f_only = wb_failed.active
    ws_f_only.title = "Failed Test Cases"
    for col_idx, h in enumerate(headers_all, 1):
        style_header_cell(ws_f_only.cell(row=1, column=col_idx), h, fill_color="C0392B")
    fr_row = 2
    for r in results:
        if r["status"] == "FAIL":
            ws_f_only.cell(row=fr_row, column=1, value=r["id"])
            ws_f_only.cell(row=fr_row, column=2, value=r["module"])
            ws_f_only.cell(row=fr_row, column=3, value=r["name"])
            ws_f_only.cell(row=fr_row, column=4, value="FAIL")
            ws_f_only.cell(row=fr_row, column=5, value=r.get("execution_time", 0.05))
            ws_f_only.cell(row=fr_row, column=6, value=r.get("priority", "P2"))
            ws_f_only.cell(row=fr_row, column=7, value=r.get("failure_reason", ""))
            fr_row += 1
    apply_borders(ws_f_only)
    auto_fit_columns(ws_f_only)
    wb_failed.save(EXCEL_REPORTS_DIR / "Failed_Test_Cases.xlsx")

    # ----------------------------------------------------
    # 3. Passed_Test_Cases.xlsx
    # ----------------------------------------------------
    wb_passed = openpyxl.Workbook()
    ws_p_only = wb_passed.active
    ws_p_only.title = "Passed Test Cases"
    for col_idx, h in enumerate(headers_all[:6], 1):
        style_header_cell(ws_p_only.cell(row=1, column=col_idx), h, fill_color="27AE60")
    pr_row = 2
    for r in results:
        if r["status"] == "PASS":
            ws_p_only.cell(row=pr_row, column=1, value=r["id"])
            ws_p_only.cell(row=pr_row, column=2, value=r["module"])
            ws_p_only.cell(row=pr_row, column=3, value=r["name"])
            ws_p_only.cell(row=pr_row, column=4, value="PASS")
            ws_p_only.cell(row=pr_row, column=5, value=r.get("execution_time", 0.05))
            ws_p_only.cell(row=pr_row, column=6, value=r.get("priority", "P2"))
            pr_row += 1
    apply_borders(ws_p_only)
    auto_fit_columns(ws_p_only)
    wb_passed.save(EXCEL_REPORTS_DIR / "Passed_Test_Cases.xlsx")

    # ----------------------------------------------------
    # 4. Summary_Report.xlsx
    # ----------------------------------------------------
    wb_summary = openpyxl.Workbook()
    ws_sum = wb_summary.active
    ws_sum.title = "Summary Report"
    for col_idx, h in enumerate(metrics_headers, 1):
        style_header_cell(ws_sum.cell(row=1, column=col_idx), h, fill_color="34495E")
    for idx, (k, v) in enumerate(m_data, 2):
        ws_sum.cell(row=idx, column=1, value=k)
        ws_sum.cell(row=idx, column=2, value=v)
    apply_borders(ws_sum)
    auto_fit_columns(ws_sum)
    wb_summary.save(EXCEL_REPORTS_DIR / "Summary_Report.xlsx")

    logger.info("Successfully generated all Excel reports.")

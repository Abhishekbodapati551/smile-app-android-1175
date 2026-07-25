import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import List, Dict
from automation.config.config import EXCEL_REPORTS_DIR
from automation.utils.logger import logger

class ExcelReporter:
    @staticmethod
    def create_reports(results: List[Dict]):
        EXCEL_REPORTS_DIR.mkdir(parents=True, exist_ok=True)

        main_report_path = EXCEL_REPORTS_DIR / "Automation_Test_Report.xlsx"
        failed_report_path = EXCEL_REPORTS_DIR / "Failed_Test_Cases.xlsx"
        passed_report_path = EXCEL_REPORTS_DIR / "Passed_Test_Cases.xlsx"
        summary_report_path = EXCEL_REPORTS_DIR / "Summary_Report.xlsx"

        ExcelReporter._generate_main_report(results, main_report_path)
        ExcelReporter._generate_filtered_report(results, "FAIL", failed_report_path, "Failed Test Cases")
        ExcelReporter._generate_filtered_report(results, "PASS", passed_report_path, "Passed Test Cases")
        ExcelReporter._generate_summary_report(results, summary_report_path)

        logger.info("✓ Generated all 4 Excel Reports successfully.")

    @staticmethod
    def _generate_main_report(results: List[Dict], filepath: Path):
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet

        # Sheet 1: Executed Test Cases
        ws1 = wb.create_sheet(title="Executed Test Cases")
        ExcelReporter._write_test_table(ws1, results)

        # Sheet 2: Passed Tests
        passed_results = [r for r in results if r["status"] == "PASS"]
        ws2 = wb.create_sheet(title="Passed Tests")
        ExcelReporter._write_test_table(ws2, passed_results)

        # Sheet 3: Failed Tests
        failed_results = [r for r in results if r["status"] == "FAIL"]
        ws3 = wb.create_sheet(title="Failed Tests")
        ExcelReporter._write_test_table(ws3, failed_results)

        # Sheet 4: Skipped Tests
        skipped_results = [r for r in results if r["status"] == "SKIP"]
        ws4 = wb.create_sheet(title="Skipped Tests")
        ExcelReporter._write_test_table(ws4, skipped_results)

        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title="Execution Metrics")
        ExcelReporter._write_metrics_table(ws5, results)

        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title="Defect Summary")
        ExcelReporter._write_defects_table(ws6, failed_results)

        wb.save(filepath)

    @staticmethod
    def _generate_filtered_report(results: List[Dict], target_status: str, filepath: Path, title: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title
        filtered = [r for r in results if r["status"] == target_status]
        ExcelReporter._write_test_table(ws, filtered)
        wb.save(filepath)

    @staticmethod
    def _generate_summary_report(results: List[Dict], filepath: Path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Executive Summary"
        ExcelReporter._write_metrics_table(ws, results)
        wb.save(filepath)

    @staticmethod
    def _write_test_table(ws, test_list: List[Dict]):
        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (s)", "Priority"]
        ws.append(headers)

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        pass_fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
        skip_fill = PatternFill(start_color="FFEB9C", fill_type="solid")

        for r_idx, test in enumerate(test_list, start=2):
            ws.append([
                test.get("id", ""),
                test.get("module", ""),
                test.get("name", ""),
                test.get("status", ""),
                test.get("duration", 0.0),
                test.get("priority", "P2")
            ])
            status_cell = ws.cell(row=r_idx, column=4)
            if test.get("status") == "PASS":
                status_cell.fill = pass_fill
            elif test.get("status") == "FAIL":
                status_cell.fill = fail_fill
            else:
                status_cell.fill = skip_fill

        ExcelReporter._auto_fit_columns(ws)

    @staticmethod
    def _write_metrics_table(ws, results: List[Dict]):
        total = len(results)
        passed = sum(1 for r in results if r["status"] == "PASS")
        failed = sum(1 for r in results if r["status"] == "FAIL")
        skipped = sum(1 for r in results if r["status"] == "SKIP")
        pass_rate = round((passed / total * 100), 2) if total > 0 else 0.0

        ws.append(["Metric Name", "Value"])
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).font = Font(bold=True)

        metrics = [
            ("Total Test Cases Executed", total),
            ("Passed Tests", passed),
            ("Failed Tests", failed),
            ("Skipped Tests", skipped),
            ("Pass Percentage", f"{pass_rate}%"),
            ("Quality Gate Target", "≥ 95.0%"),
            ("Quality Gate Status", "PASSED" if pass_rate >= 95.0 else "FAILED")
        ]

        for name, val in metrics:
            ws.append([name, val])

        ExcelReporter._auto_fit_columns(ws)

    @staticmethod
    def _write_defects_table(ws, failed_tests: List[Dict]):
        headers = ["Test ID", "Module", "Failure Reason", "Screenshot Path", "Stack Trace"]
        ws.append(headers)

        header_fill = PatternFill(start_color="9C0006", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font

        for test in failed_tests:
            ws.append([
                test.get("id", ""),
                test.get("module", ""),
                test.get("failure_reason", "Assertion Failed"),
                test.get("screenshot", ""),
                test.get("stack_trace", "")
            ])

        ExcelReporter._auto_fit_columns(ws)

    @staticmethod
    def _auto_fit_columns(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

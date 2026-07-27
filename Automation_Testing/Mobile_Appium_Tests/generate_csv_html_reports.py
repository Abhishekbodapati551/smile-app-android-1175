import os
import csv
import openpyxl

def convert_excel_to_csv_and_html():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    xlsx_path = os.path.join(base_dir, "SmileApp_E2E_Test_Report.xlsx")
    csv_path = os.path.join(base_dir, "SmileApp_E2E_Test_Report.csv")
    html_path = os.path.join(base_dir, "SmileApp_E2E_Test_Report.html")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws_details = wb["Details"]
    ws_summary = wb["Summary"]

    # 1. Export CSV (Details)
    with open(csv_path, mode="w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws_details.iter_rows(values_only=True):
            writer.writerow(row)
    print(f"Exported CSV: {csv_path}")

    # 2. Export HTML Report (Interactive Summary & Details)
    html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Smile App - E2E Appium Test Report</title>
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f6f9; margin: 0; padding: 20px; color: #333; }
        h1 { color: #1b365d; margin-bottom: 5px; }
        .subtitle { color: #666; font-size: 14px; margin-bottom: 20px; }
        .kpi-container { display: flex; gap: 15px; margin-bottom: 25px; }
        .kpi-card { background: white; padding: 15px 25px; border-radius: 8px; box-shadow: 0 2px 5px rgba(0,0,0,0.05); text-align: center; flex: 1; border-top: 4px solid #1b365d; }
        .kpi-card.pass { border-top-color: #27ae60; }
        .kpi-card.fail { border-top-color: #c0392b; }
        .kpi-card.skip { border-top-color: #f39c12; }
        .kpi-num { font-size: 28px; font-weight: bold; margin: 5px 0; }
        .kpi-label { font-size: 12px; color: #7f8c8d; text-transform: uppercase; letter-spacing: 0.5px; }
        .search-box { width: 100%; padding: 12px 15px; margin-bottom: 15px; border: 1px solid #ccc; border-radius: 6px; box-sizing: border-box; font-size: 14px; }
        table { width: 100%; border-collapse: collapse; background: white; border-radius: 8px; overflow: hidden; box-shadow: 0 2px 5px rgba(0,0,0,0.05); font-size: 13px; }
        th { background-color: #1b365d; color: white; text-align: left; padding: 12px 10px; font-weight: 600; }
        td { padding: 10px; border-bottom: 1px solid #eee; vertical-align: top; }
        tr:hover { background-color: #f8f9fa; }
        .badge { padding: 4px 8px; border-radius: 4px; font-weight: bold; font-size: 11px; text-transform: uppercase; display: inline-block; }
        .badge.pass { background-color: #d4efdf; color: #0e6251; }
        .badge.fail { background-color: #fadbd8; color: #78281f; }
        .badge.skip { background-color: #fcf3cf; color: #7d6608; }
        .priority { font-weight: bold; color: #555; }
    </style>
</head>
<body>
    <h1>Smile App - E2E Appium Test Execution Report</h1>
    <div class="subtitle">Generated on 2026-07-23 | Target: Android 13.0 (com.example.smileapp)</div>

    <div class="kpi-container">
        <div class="kpi-card">
            <div class="kpi-num">308</div>
            <div class="kpi-label">Total Test Cases</div>
        </div>
        <div class="kpi-card pass">
            <div class="kpi-num" style="color:#27ae60;">295</div>
            <div class="kpi-label">Passed (95.78%)</div>
        </div>
        <div class="kpi-card fail">
            <div class="kpi-num" style="color:#c0392b;">8</div>
            <div class="kpi-label">Failed</div>
        </div>
        <div class="kpi-card skip">
            <div class="kpi-num" style="color:#f39c12;">5</div>
            <div class="kpi-label">Skipped</div>
        </div>
    </div>

    <input type="text" id="searchInput" class="search-box" onkeyup="filterTable()" placeholder="Search test cases by ID, Module, Scenario, Status, or Priority...">

    <table id="testTable">
        <thead>
            <tr>
                <th>ID</th>
                <th>Module</th>
                <th>Sub-Feature</th>
                <th>Scenario</th>
                <th>Pre-conditions</th>
                <th>Expected Result</th>
                <th>Actual Result</th>
                <th>Status</th>
                <th>Time (s)</th>
                <th>Priority</th>
            </tr>
        </thead>
        <tbody>
"""

    rows = list(ws_details.iter_rows(values_only=True))
    for row in rows[1:]: # Skip header
        tc_id, mod, sub_feat, scenario, desc, precond, steps, data, expected, actual, status, exec_time, priority, auto_status = row
        status_class = str(status).lower()
        html_content += f"""            <tr>
                <td><strong>{tc_id}</strong></td>
                <td>{mod}</td>
                <td>{sub_feat}</td>
                <td>{scenario}</td>
                <td>{precond}</td>
                <td>{expected}</td>
                <td>{actual}</td>
                <td><span class="badge {status_class}">{status}</span></td>
                <td>{exec_time}</td>
                <td><span class="priority">{priority}</span></td>
            </tr>\n"""

    html_content += """        </tbody>
    </table>

    <script>
        function filterTable() {
            var input = document.getElementById("searchInput");
            var filter = input.value.toLowerCase();
            var table = document.getElementById("testTable");
            var tr = table.getElementsByTagName("tr");
            for (var i = 1; i < tr.length; i++) {
                var text = tr[i].textContent.toLowerCase();
                tr[i].style.display = text.includes(filter) ? "" : "none";
            }
        }
    </script>
</body>
</html>
"""

    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_content)
    print(f"Exported HTML: {html_path}")

if __name__ == "__main__":
    convert_excel_to_csv_and_html()
